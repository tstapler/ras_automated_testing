from __future__ import division
from rocprotocol import RocConnection, OpcodeLib
from rocprotocol.rocdata import get_data
from time import sleep
from collections import Counter
from openpyxl import Workbook
import struct
import visa
import click
from time import time
from math import log
import configparser

config = configparser.ConfigParser()
config.read('config.ini')
defaults = config['DEFAULT']
test_params = config['TEST_PARAM']

readings = []

calibrations = [12, 20, 26] #Voltages to calibrate at

miliamps = [(4,0.0), (12,50.0), (20,100.0)] #As a percentage of 20mA

#Create a list of the bytes representing the floats in miliamps, only in int form
#EG:for milamps = [0.0, 50.0, 100.0], output = [[0,0,0,0],[0,0,72,66],[0,0,200,66]]
miliamps = [(y[0],[ x for x in struct.pack('f', y[1])]) for y in miliamps]

lib = OpcodeLib() #Library of Opcodes as defined in .json file


def take_reading(rtu, minimum=5, maximum=12):
    measuring = True
    AD = []
    EU = []
    while measuring:
        #Data format for 180: Number of Requested, T, L, P
        rtu.send_opcode(180,opcode=lib.opcode(180,data=[1,3,ai_input_logical,17]))
        val = get_data(rtu.read_response(),dtype="int")
        AD.append(val)
        rtu.send_opcode(180,opcode=lib.opcode(180,data=[1,3,ai_input_logical,14]))
        val = get_data(rtu.read_response(),dtype="float")
        EU.append(val)
        if len(AD) > minimum :
            ADcounts = Counter(AD)
            ADcommon = ADcounts.most_common()
            if ADcommon[0][1]/sum(ADcounts) * 100 > .70:
                #If there is a clear choice log it
                measuring = False
                return ADcommon[0][0], EU[AD.index(ADcommon[0][0])]
            elif len(AD) >= maximum:
                measuring = False
                return ADcommon[0][0], EU[AD.index(ADcommon[0][0])]

def calibrate(rtu, input_rtu, powersupply, voltage=20):
    powersupply.write('APPLy ' + str(voltage))
    # 0%
    #Set input to 4mA
    input_rtu.send_opcode(181,opcode = lib.opcode(181,data=[1,4,ai_output_logical,6] + miliamps[0][1]))
    input_rtu.read_response()
    #Take current reading

    sleep(int(test_params['calibrate_soak']))
    ADreading, EUreading = take_reading(rtu,maximum=30)
    zero_point = ADreading
    #Convert the readings into a uint16 represented by a list of 4 integers
    ADreading = [x for x in struct.pack("H",ADreading)]
    #Set adjusted A/D 0%
    rtu.send_opcode(181,opcode=lib.opcode(181,data=[1,3,ai_input_logical,4] + ADreading))
    rtu.read_response()

    # 100%
    #Set input to 20mA
    input_rtu.send_opcode(181,opcode = lib.opcode(181,data=[1,4,ai_output_logical,6] + miliamps[2][1]))
    input_rtu.read_response()
    sleep(int(test_params['calibrate_soak']))
    #Take current reading
    ADreading, EUreading = take_reading(rtu,maximum=30)
    hundred_point = ADreading
    #Convert the readings into a uint16 represented by a list of 4 integers
    ADreading = [x for x in struct.pack("H",ADreading)]
    #Set adjusted A/D 100%
    rtu.send_opcode(181,opcode=lib.opcode(181,data=[1,3,ai_input_logical,5] + ADreading))
    rtu.read_response()


    # 50%
    #Set input to 12mA
    input_rtu.send_opcode(181,opcode = lib.opcode(181,data=[1,4,ai_output_logical,6] + miliamps[1][1]))
    input_rtu.read_response()
    sleep(int(test_params['calibrate_soak']))
    #Take current reading
    ADreading, EUreading = take_reading(rtu,maximum=30)
    fifty_point = ADreading
    #Convert the readings into a uint16 represented by a list of 4 integers
    ADreading = [x for x in struct.pack("H",ADreading)]
    #Set Mid Point Raw 1
    rtu.send_opcode(181,opcode=lib.opcode(181,data=[1,3,ai_input_logical,21] + ADreading))
    rtu.read_response()
    #Set Mid Point EU 1
    rtu.send_opcode(181,opcode=lib.opcode(181,data=[1,3,ai_input_logical,26] + miliamps[1][1]))
    rtu.read_response

    return zero_point, fifty_point, hundred_point

def xfrange(start, stop, step):
    old_start = start #backup this value

    digits = int(round(log(10000, 10)))+1 #get number of digits
    magnitude = 10**digits
    stop = int(magnitude * stop) #convert from
    step = int(magnitude * step) #0.1 to 10 (e.g.)

    if start == 0:
        start = 10**(digits-1)
    else:
        start = 10**(digits)*start

    data = []   #create array

    #calc number of iterations
    end_loop = int((stop-start)//step)
    if old_start == 0:
        end_loop += 1

    acc = start

    for i in range(0, end_loop):
        data.append(acc/magnitude)
        acc += step

    return data

def set_headers(worksheet):
    headers = ["Input(mA)", "Voltage(V)", "Reading", "A/D count", "Counts Off"]
    for row_numb in [1, 20, 39]:
        for row in worksheet.iter_rows('A' + str(row_numb) + ':' + 'E' + str(row_numb)):
            for cell, header in zip(row, headers):
                cell.value = header
    worksheet['F1'].value = "0 Point"
    worksheet['F20'].value = "50 Point"
    worksheet['F39'].value = "100 Point"

def main(software_version="1 70h", board_serial = "w38334x0012AK1432180",
        ai="ai1", temperature="21", ai_in_logical=33, ai_out_logical=11, fb_com_port="COM9", roc_com_port="COM10"):

    global ai_input_logical
    ai_input_logical = ai_in_logical
    global ai_output_logical
    ai_output_logical = ai_out_logical

    #Set up excel workbook
    results_sheet = Workbook()
    results_sheet.remove_sheet(results_sheet.worksheets[0])
    dest_filename = 'AIdrift_' + ai + '_'+ software_version + "_" + board_serial + "_" + temperature

    print("\n\n\n\n----Starting AI Drift Test----")
    print("Board Serial:", board_serial)
    print("Firmware Version:", software_version)
    print("AI Under Test:", ai)
    print("Test Temperature:", temperature)
    start = time()
    #Initialize connection to 107
    fb107 = RocConnection(baud=19200, port=fb_com_port, timeout=0.05)

    #Initialize connection to 312
    roc312 = RocConnection(baud=9600, port=roc_com_port)

    #Initialize connection to Agilent E3634A
    rm = visa.ResourceManager()
    powersupply = rm.open_resource('ASRL1::INSTR') #ASRL1 == COM1

    for calibration in calibrations:
        #TODO: Create worksheets earlier
        print("\n\nStarting", calibration, "Volt", "Calibration:")
        calib_points = {}

        calib_points['4'], calib_points['12'], calib_points['20'] = calibrate(fb107, roc312, powersupply, voltage=calibration)

        sheet = results_sheet.create_sheet(title=str(calibration) + "V Calibration")
        #Set All of the Header Rows on each sheet
        set_headers(sheet)
        #Set the 0 point
        sheet['F2'].value = calib_points['4']
        #Set the 50 point
        sheet['F21'].value = calib_points['12']
        #Set the 100 point
        sheet['F40'].value = calib_points['20']

        row_on_sheet = 2
        for decimal, amperage in miliamps:
            data = [1,4,ai_output_logical,6] + amperage
            roc312.send_opcode(181,opcode=lib.opcode(181,data=data))
            curr_calib_point = calib_points[str(decimal)]
            sleep(1)
            first = True
            with click.progressbar(xfrange(calibration-4, calibration+4.5,.5), label=str(calibration) + "V calibration -> " + str(decimal) + 'mA') as voltages:
                for voltage in voltages:
                    powersupply.write('APPLy ' + str(voltage))
		    #The initial soak value for the first voltage in each range
                    if first:
                        sleep(int(test_params['initial_soak']))
                        first = False
		    #The intermediate wait between voltages
                    else:
                        sleep(int(test_params['intermediate_soak']))
                    ADval, EUval = take_reading(fb107, minimum=int(test_params['minimum_wait']),maximum=int(test_params['maximum_wait']))
                    for row in sheet.iter_rows('A' + str(row_on_sheet) + ':' + 'E' + str(row_on_sheet)):
                        for cell, data in zip(row, [decimal, voltage, EUval, ADval, ADval - curr_calib_point]):
                            cell.value = data
                    readings.append((decimal, voltage, EUval, ADval, ADval - curr_calib_point))
                    row_on_sheet += 1
            row_on_sheet += 2
            try:
                results_sheet.save('Results\\' + dest_filename + '.xlsx')
            except:
                results_sheet.save("ai_drift_permission_error.xlsx")


    try:
        results_sheet.save('Results\\' + dest_filename + '.xlsx')
    except:
        results_sheet.save("ai_drift_permission_error.xlsx")

    end = time()
    print("\n\nTotal Time Elapsed:", end - start, "\n\n\n")

@click.command()
@click.option('-v', prompt="Firmware Version", default=defaults['version'], help='Firmware version being tested')
@click.option('--board_type', prompt="Board Type (CPU/6PT)", default=defaults['board'], help='Type of board CPU or 6PT')
@click.option('-s', prompt="Serial Number", default=defaults['serial'], help='Serial number of the board being tested (Numbers after W38334x0012)')
@click.option('-ai', prompt="Input being tested (AI1/AI2)", default=defaults['input'], help="Which AI is being tested")
@click.option('-t', prompt="Temperature", default=defaults['temperature'], help="Current temperature in degrees celsius")
<<<<<<< HEAD
def cli(v, board_type, s, ai, t):
=======
def cli(v, board_type, s, ai, t, com_port_107, com_port_312):
>>>>>>> 3f487024b8f68095b2f8def9db9c376e07f6f900
    if board_type == "6PT":
        b_type = "W38334x0012"
    else:
        b_type = "W48084x0012"

    if board_type is "CPU":
        if ai is "AI1":
            in_logical = 8
        else:
            in_logical = 9
    elif board_type is "6PT":
        if ai is "AI1":
            in_logical = 32
        else:
            in_logical = 33
    else:
        print("Invalid Board type or Input")

    main(software_version=v, board_serial=b_type + s, ai=ai, temperature=t,
            ai_in_logical=in_logical, ai_out_logical=defaults['ao_logical'], fb_com_port=defaults['com_port_107'], roc_com_port=defaults['com_port_312'])

    again = click.prompt("Would you like to run another test?", type=bool)

    if again:
        cli()


if __name__ == "__main__":
    cli()
